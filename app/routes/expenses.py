import io
import json
from decimal import Decimal, ROUND_HALF_UP
from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, jsonify
from sqlalchemy import func
from app import db
from app.models import User, Expense, InstallmentGroup, RecurringGroup, CreditCard, CategoryRule
from app.forms import ExpenseForm
from app.services.categorizer import categorize_batch, create_rule_if_missing, is_duplicate
from app.utils import tenant_users, tenant_user_ids, MONTH_NAMES_SHORT, month_offset, user_color_map, get_month_year, _brl
from datetime import datetime

expenses_bp = Blueprint('expenses', __name__, url_prefix='/expenses')

ITEMS_PER_PAGE = 20

CATEGORIES = [
    'Alimentação', 'Transporte', 'Saúde', 'Lazer', 'Moradia',
    'Educação', 'Vestuário', 'Serviços', 'Compras', 'Outros',
]

_BANK_PAYMENTS = {'PIX', 'Cartão de Débito', 'Cartão de Crédito'}


def _bank_from_form(form, payment):
    return (form.bank.data or None) if payment in _BANK_PAYMENTS else None


def _tenant_categories(uids):
    db_cats = {c[0] for c in db.session.query(Expense.category)
               .filter(Expense.user_id.in_(uids)).distinct()}
    return sorted(set(CATEGORIES) | db_cats)


def _suggest_category(description, uids):
    term = f'%{description.lower()}%'
    row = (db.session.query(Expense.category, func.count('*').label('cnt'))
           .filter(Expense.user_id.in_(uids),
                   func.lower(Expense.description).like(term))
           .group_by(Expense.category)
           .order_by(func.count('*').desc())
           .first())
    return row[0] if row else None


def _parse_c6pdf(file_bytes, form):
    from app.importers.c6 import parse_c6_pdf
    pwd = form.get('pdf_password', '').strip() or None
    return parse_c6_pdf(file_bytes, password=pwd), 'C6'


def _parse_ofx(file_bytes, form):
    from app.importers.ofx import parse_ofx
    return parse_ofx(file_bytes), None


def _parse_nubank(file_bytes, form):
    from app.importers.nubank import parse_nubank_csv
    return parse_nubank_csv(file_bytes), 'Nubank'


def _parse_csv_generic(file_bytes, form):
    from app.importers.csv_generic import parse_csv_generic
    return parse_csv_generic(file_bytes), None


_PARSERS = {
    'c6pdf':       _parse_c6pdf,
    'ofx':         _parse_ofx,
    'nubank_csv':  _parse_nubank,
    'csv_generic': _parse_csv_generic,
}


def _safe_next(url, fallback):
    """Valida URL de redirect: permite apenas URLs relativas sem esquema ou domínio."""
    if not url or not url.startswith('/') or url.startswith('//'):
        return fallback
    return url


@expenses_bp.route('/')
def index():
    users = tenant_users().order_by(User.name).all()
    uids = [u.id for u in users]
    now = datetime.now()

    user_id = request.args.get('user_id', type=int)
    month, year = get_month_year()
    if request.args.get('month', type=int) == 0:
        month = 0  # 0 = todos os meses
    category = request.args.get('category', '')
    payment_method = request.args.get('payment_method', '')
    bank_filter = request.args.get('bank', '')
    paid_filter = request.args.get('paid', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = request.args.get('page', 1, type=int)

    date_int = Expense.year * 10000 + Expense.month * 100 + Expense.day
    query = Expense.query.filter(Expense.user_id.in_(uids))

    if date_from or date_to:
        # intervalo de datas explícito substitui o filtro mês/ano
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(date_int >= df.year * 10000 + df.month * 100 + df.day)
            except ValueError:
                date_from = ''
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d')
                query = query.filter(date_int <= dt.year * 10000 + dt.month * 100 + dt.day)
            except ValueError:
                date_to = ''
    else:
        query = query.filter(Expense.year == year)
        if month != 0:
            query = query.filter(Expense.month == month)

    if user_id and user_id in uids:
        query = query.filter(Expense.user_id == user_id)
    if category:
        query = query.filter(Expense.category == category)
    if payment_method:
        query = query.filter(Expense.payment_method == payment_method)
    if bank_filter:
        query = query.filter(Expense.bank == bank_filter)
    if paid_filter == 'pago':
        query = query.filter(Expense.paid == True)
    elif paid_filter == 'pendente':
        query = query.filter(Expense.paid.isnot(True))

    query = query.order_by(Expense.month.desc(), Expense.day.desc(), Expense.created_at.desc())
    pagination = query.paginate(page=page, per_page=ITEMS_PER_PAGE, error_out=False)
    expenses = pagination.items

    total_mes = sum(float(e.amount) for e in query.all())

    categories = (db.session.query(Expense.category)
                  .filter(Expense.user_id.in_(uids))
                  .distinct().order_by(Expense.category).all())
    categories = [c[0] for c in categories]

    payment_methods = ['PIX', 'Cartão de Débito', 'Cartão de Crédito', 'Dinheiro', 'VR', 'VA']

    banks = (db.session.query(Expense.bank)
             .filter(Expense.user_id.in_(uids),
                     Expense.bank.isnot(None), Expense.bank != '')
             .distinct().order_by(Expense.bank).all())
    banks = [b[0] for b in banks]

    return render_template('expenses/list.html',
                           expenses=expenses,
                           pagination=pagination,
                           users=users,
                           categories=categories,
                           payment_methods=payment_methods,
                           banks=banks,
                           total_mes=total_mes,
                           today=now.date(),
                           user_colors=user_color_map(users),
                           filters={'user_id': user_id, 'month': month, 'year': year,
                                    'category': category, 'payment_method': payment_method,
                                    'bank': bank_filter, 'paid': paid_filter,
                                    'date_from': date_from, 'date_to': date_to})


_EXPORT_HEADERS = [
    'Pessoa', 'Dia', 'Mês', 'Ano', 'Descrição', 'Categoria',
    'Forma de Pagamento', 'Banco', 'Valor (R$)', 'Parcela', 'Recorrente',
]


def _expense_rows(expenses):
    rows = []
    for e in expenses:
        parcela = ''
        if e.installment_group_id:
            parcela = f'{e.installment_number}/{e.installment_group.num_installments}'
        elif e.recurring_group_id:
            parcela = f'Recorrente {e.recurring_number}/{e.recurring_group.num_recurrences}'
        rows.append([
            e.user.name, e.day, e.month, e.year,
            e.description, e.category, e.payment_method, e.bank or '',
            float(e.amount), parcela,
            'Sim' if e.recurring_group_id else 'Não',
        ])
    return rows


def _make_xlsx(rows, year, month):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Despesas'
    ws.append(_EXPORT_HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0D6EFD')
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        display = list(row)
        display[8] = _brl(row[8])
        ws.append(display)

    col_widths = [14, 6, 6, 7, 42, 16, 20, 16, 16, 13, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f'despesas_{year}_{month:02d}.xlsx'
    return Response(
        out.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _make_pdf(rows, year, month):
    from fpdf import FPDF

    period = 'Todos os meses' if month == 0 else f'{month:02d}/{year}'

    class _PDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 13)
            self.cell(0, 10, f'Despesas - {period}', align='C', new_x='LMARGIN', new_y='NEXT')
            self.ln(3)

        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Pagina {self.page_no()}', align='C')

    pdf = _PDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    cols = [
        ('Pessoa', 22), ('Dia', 9), ('Mes', 9), ('Ano', 12),
        ('Descricao', 63), ('Categoria', 24), ('Pagamento', 28),
        ('Banco', 22), ('Valor (R$)', 24), ('Parcela', 18), ('Recorrente', 14),
    ]
    row_h = 7

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(13, 110, 253)
    pdf.set_text_color(255, 255, 255)
    for label, w in cols:
        pdf.cell(w, row_h, label, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('Helvetica', size=8)
    pdf.set_text_color(0, 0, 0)
    for i, row in enumerate(rows):
        r, g, b = (240, 244, 255) if i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(r, g, b)
        display = list(row)
        display[8] = _brl(row[8])
        for j, (_, w) in enumerate(cols):
            pdf.cell(w, row_h, str(display[j]), border=1, fill=True,
                     align='R' if j == 8 else 'L')
        pdf.ln()

    out = io.BytesIO()
    pdf.output(out)
    out.seek(0)
    filename = f'despesas_{year}_{month:02d}.pdf'
    return Response(
        out.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@expenses_bp.route('/export')
def export_csv():
    now = datetime.now()
    uids = tenant_user_ids()
    user_id = request.args.get('user_id', type=int)
    month = request.args.get('month', now.month, type=int)
    year = request.args.get('year', now.year, type=int)
    category = request.args.get('category', '')
    fmt = request.args.get('fmt', 'xlsx')

    if month == 0:
        query = Expense.query.filter(Expense.user_id.in_(uids), Expense.year == year)
    else:
        query = Expense.query.filter(Expense.user_id.in_(uids), Expense.year == year,
                                     Expense.month == month)
    if user_id and user_id in uids:
        query = query.filter(Expense.user_id == user_id)
    if category:
        query = query.filter(Expense.category == category)
    query = query.order_by(Expense.day.asc(), Expense.created_at.asc())

    rows = _expense_rows(query.all())
    if fmt == 'pdf':
        return _make_pdf(rows, year, month)
    return _make_xlsx(rows, year, month)


def _resolve_card(form, payment, uids):
    """
    Retorna (card, card_id, billing_month, billing_year).

    - Crédito: ajusta mês de fatura pelo best_buy_day
    - Débito:  usa cartão selecionado (sem ajuste de mês)
    - VR/VA:   usa cartão selecionado OU auto-associa ao cartão VR/VA do usuário
    """
    raw_id = form.card_id.data or 0
    month  = form.month.data
    year   = form.year.data

    if payment == 'Cartão de Crédito':
        if not raw_id:
            return None, None, month, year
        card = CreditCard.query.filter(CreditCard.id == raw_id,
                                       CreditCard.user_id.in_(uids)).first()
        if not card:
            return None, None, month, year
        bm, by = month_offset(month, year,
                              1 if form.day.data > card.best_buy_day else 0)
        return card, card.id, bm, by

    if payment in ('Cartão de Débito', 'VR', 'VA'):
        card = None
        # 1) Cartão selecionado explicitamente no modal
        if raw_id:
            card = CreditCard.query.filter(CreditCard.id == raw_id,
                                           CreditCard.user_id.in_(uids)).first()
        # 2) VR/VA: auto-associar com o cartão do tipo correto se não informado
        if not card and payment in ('VR', 'VA'):
            ct = 'vr' if payment == 'VR' else 'va'
            card = (CreditCard.query
                    .filter(CreditCard.user_id.in_(uids),
                            CreditCard.card_type == ct)
                    .first())
        return (card, card.id if card else None, month, year)

    return None, None, month, year


@expenses_bp.route('/add', methods=['GET', 'POST'])
def add():
    import json
    users = tenant_users().order_by(User.name).all()
    uids = [u.id for u in users]
    form = ExpenseForm()
    form.user_id.choices = [(u.id, u.name) for u in users]

    credit_cards = CreditCard.query.filter(CreditCard.user_id.in_(uids)).order_by(CreditCard.bank).all()
    form.card_id.choices = [(0, '— Nenhum (usar banco acima) —')] + [
        (c.id, c.label) for c in credit_cards
    ]

    now = datetime.now()
    if request.method == 'GET':
        form.year.data = now.year
        form.month.data = now.month
        form.day.data = now.day
        form.num_installments.data = 2
        form.recurring_times.data = 12
        form.card_id.data = 0

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if form.validate_on_submit():
        payment = form.payment_method.data
        bank = _bank_from_form(form, payment)

        purchase_month, purchase_year = form.month.data, form.year.data

        card, card_id, billing_month, billing_year = _resolve_card(form, payment, uids)
        if card:
            bank = card.bank or bank
            form.month.data = billing_month
            form.year.data = billing_year

        is_parcelado = (payment == 'Cartão de Crédito' and
                        form.credit_type.data == 'parcelado')
        is_recurring = form.is_recurring.data

        if is_parcelado:
            _create_installments(form, bank, card_id=card_id,
                                 purchase_month=purchase_month, purchase_year=purchase_year)
        elif is_recurring:
            _create_recurring(form, bank, payment, card_id=card_id,
                              purchase_month=purchase_month, purchase_year=purchase_year)
        else:
            dup = is_duplicate(
                form.user_id.data,
                form.year.data, form.month.data, form.day.data,
                float(form.amount.data), form.description.data,
            )
            expense = Expense(
                user_id=form.user_id.data,
                description=form.description.data,
                amount=form.amount.data,
                category=form.category.data,
                payment_method=payment,
                bank=bank,
                year=form.year.data,
                month=form.month.data,
                day=form.day.data,
                purchase_month=purchase_month,
                purchase_year=purchase_year,
                card_id=card_id,
                paid=bool(form.paid.data),
            )
            db.session.add(expense)
            db.session.commit()

            if dup:
                msg = '⚠️ Despesa salva, mas parece duplicada — verifique se já existe igual.'
                if is_ajax:
                    return jsonify({'status': 'ok', 'message': msg, 'duplicate': True})
                flash(msg, 'warning')
                return redirect(url_for('expenses.index'))

        if is_ajax:
            return jsonify({'status': 'ok', 'message': 'Despesa adicionada com sucesso!'})

        flash('Despesa adicionada com sucesso!', 'success')
        return redirect(url_for('expenses.index'))

    # Erros de validação
    if is_ajax:
        errors = {f.name: f.errors[0] for f in form if f.errors}
        return jsonify({'status': 'error', 'errors': errors}), 422

    card_data_json = json.dumps({
        c.id: {
            'due_day':         c.due_day,
            'best_buy_day':    c.best_buy_day,
            'bank':            c.bank or '',
            'label':           c.label,
            'credit_limit':    float(c.credit_limit) if c.credit_limit else None,
            'card_type':       c.card_type or '',
            'supports_credit': bool(c.supports_credit),
            'supports_debit':  bool(c.supports_debit),
        }
        for c in credit_cards
    })
    return render_template('expenses/add.html', form=form, users=users,
                           all_categories=_tenant_categories(uids),
                           credit_cards=credit_cards,
                           card_data_json=card_data_json)


@expenses_bp.route('/edit/<int:expense_id>', methods=['GET', 'POST'])
def edit(expense_id):
    uids = tenant_user_ids()
    expense = Expense.query.filter(Expense.id == expense_id, Expense.user_id.in_(uids)).first_or_404()
    users = tenant_users().order_by(User.name).all()
    uids  = [u.id for u in users]

    credit_cards = CreditCard.query.filter(CreditCard.user_id.in_(uids)).order_by(CreditCard.bank).all()

    form = ExpenseForm(obj=expense)
    form.user_id.choices = [(u.id, u.name) for u in users]
    form.card_id.choices = [(0, '— Nenhum (usar banco acima) —')] + [
        (c.id, c.label) for c in credit_cards
    ]

    if request.method == 'GET':
        form.credit_type.data = 'avista'
        form.card_id.data = expense.card_id or 0

    if form.validate_on_submit():
        payment     = form.payment_method.data
        card_id_val = form.card_id.data
        new_card_id = card_id_val if (card_id_val and card_id_val > 0) else None

        expense.user_id        = form.user_id.data
        expense.description    = form.description.data
        expense.amount         = form.amount.data
        expense.category       = form.category.data
        expense.payment_method = payment
        expense.bank           = _bank_from_form(form, payment)
        expense.year           = form.year.data
        expense.month          = form.month.data
        expense.day            = form.day.data
        expense.paid           = form.paid.data
        expense.card_id        = new_card_id

        apply_all = request.form.get('apply_all') == '1'

        if apply_all and (expense.installment_group_id or expense.recurring_group_id):
            # Campos propagados para todas as parcelas/recorrências
            shared = {
                'description':    expense.description,
                'category':       expense.category,
                'payment_method': expense.payment_method,
                'bank':           expense.bank,
                'card_id':        expense.card_id,
            }

            if expense.installment_group_id:
                siblings = Expense.query.filter(
                    Expense.installment_group_id == expense.installment_group_id,
                    Expense.id != expense.id,
                    Expense.user_id.in_(uids),
                ).all()
                for s in siblings:
                    for field, val in shared.items():
                        setattr(s, field, val)
                total = len(siblings) + 1
                flash(f'Alterações aplicadas a todas as {total} parcelas.', 'success')

            elif expense.recurring_group_id:
                siblings = Expense.query.filter(
                    Expense.recurring_group_id == expense.recurring_group_id,
                    Expense.id != expense.id,
                    Expense.user_id.in_(uids),
                ).all()
                for s in siblings:
                    for field, val in shared.items():
                        setattr(s, field, val)
                    s.amount = expense.amount  # recorrência tem mesmo valor
                total = len(siblings) + 1
                flash(f'Alterações aplicadas a todas as {total} recorrências.', 'success')
        else:
            flash('Despesa atualizada com sucesso!', 'success')

        db.session.commit()
        next_url = _safe_next(request.form.get('next'), url_for('expenses.index'))
        return redirect(next_url)

    card_data_json = json.dumps({
        c.id: {
            'due_day':         c.due_day,
            'best_buy_day':    c.best_buy_day,
            'bank':            c.bank or '',
            'label':           c.label,
            'credit_limit':    float(c.credit_limit) if c.credit_limit else None,
            'card_type':       c.card_type or '',
            'supports_credit': bool(c.supports_credit),
            'supports_debit':  bool(c.supports_debit),
        }
        for c in credit_cards
    })
    return render_template('expenses/edit.html', form=form, expense=expense,
                           users=users, all_categories=_tenant_categories(uids),
                           credit_cards=credit_cards, card_data_json=card_data_json)


@expenses_bp.route('/delete/<int:expense_id>', methods=['POST'])
def delete(expense_id):
    uids = tenant_user_ids()
    expense = Expense.query.filter(Expense.id == expense_id, Expense.user_id.in_(uids)).first_or_404()
    db.session.delete(expense)
    db.session.commit()
    flash('Despesa eliminada.', 'warning')
    return redirect(url_for('expenses.index'))


@expenses_bp.route('/delete-group/<int:group_id>', methods=['POST'])
def delete_group(group_id):
    uids = tenant_user_ids()
    group = InstallmentGroup.query.filter(InstallmentGroup.id == group_id, InstallmentGroup.user_id.in_(uids)).first_or_404()
    count = group.installments.count()
    db.session.delete(group)
    db.session.commit()
    flash(f'Todas as {count} parcelas foram eliminadas.', 'warning')
    return redirect(url_for('expenses.index'))


@expenses_bp.route('/toggle-paid/<int:expense_id>', methods=['POST'])
def toggle_paid(expense_id):
    uids = tenant_user_ids()
    expense = Expense.query.filter(Expense.id == expense_id, Expense.user_id.in_(uids)).first_or_404()
    expense.paid = not bool(expense.paid)
    db.session.commit()
    return redirect(_safe_next(request.form.get('next'), url_for('expenses.index')))


@expenses_bp.route('/bulk-paid', methods=['POST'])
def bulk_paid():
    uids = tenant_user_ids()
    ids = request.form.getlist('expense_ids')
    count = 0
    for eid in ids:
        try:
            expense = Expense.query.filter(
                Expense.id == int(eid),
                Expense.user_id.in_(uids)
            ).first()
            if expense and not expense.paid:
                expense.paid = True
                count += 1
        except (ValueError, TypeError):
            pass
    db.session.commit()
    flash(f'{count} despesa(s) marcada(s) como pagas!', 'success')
    return redirect(_safe_next(request.form.get('next'), url_for('main.index')))


@expenses_bp.route('/delete-recurring/<int:group_id>', methods=['POST'])
def delete_recurring(group_id):
    uids = tenant_user_ids()
    group = RecurringGroup.query.filter(RecurringGroup.id == group_id, RecurringGroup.user_id.in_(uids)).first_or_404()
    count = group.recurrences.count()
    db.session.delete(group)
    db.session.commit()
    flash(f'Todas as {count} recorrências foram eliminadas.', 'warning')
    return redirect(url_for('expenses.index'))


@expenses_bp.route('/import', methods=['GET'])
def import_bank():
    users = tenant_users().order_by(User.name).all()
    return render_template('expenses/import_upload.html', users=users)


@expenses_bp.route('/import/finfam', methods=['POST'])
def import_finfam():
    import openpyxl

    upload = request.files.get('file')
    if not upload or not upload.filename:
        flash('Selecione um arquivo .xlsx.', 'danger')
        return redirect(url_for('expenses.import_bank'))
    if not upload.filename.lower().endswith('.xlsx'):
        flash('O arquivo deve ser um .xlsx exportado pelo FinFam.', 'danger')
        return redirect(url_for('expenses.import_bank'))

    users = tenant_users().order_by(User.name).all()
    user_map = {u.name.strip().lower(): u for u in users}

    try:
        wb = openpyxl.load_workbook(filename=upload.stream, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        flash(f'Erro ao ler o arquivo: {e}', 'danger')
        return redirect(url_for('expenses.import_bank'))

    imported = skipped = 0
    for row in rows:
        if not row or not any(row):
            continue
        try:
            nome, dia, mes, ano, descricao, categoria, pagamento, banco, valor_raw = (row + (None,) * 9)[:9]

            user = user_map.get(str(nome or '').strip().lower())
            if not user:
                skipped += 1
                continue

            if isinstance(valor_raw, (int, float)):
                amount = float(valor_raw)
            else:
                s = str(valor_raw).replace('R$', '').strip().replace('.', '').replace(',', '.')
                amount = float(s)

            if amount <= 0:
                skipped += 1
                continue

            dia, mes, ano = int(dia), int(mes), int(ano)
            if not (1 <= mes <= 12 and 2000 <= ano <= 2100 and 1 <= dia <= 31):
                skipped += 1
                continue

            db.session.add(Expense(
                user_id=user.id,
                description=str(descricao or '').strip()[:200],
                amount=amount,
                category=str(categoria or 'Outros').strip(),
                payment_method=str(pagamento or 'PIX').strip(),
                bank=str(banco or '').strip() or None,
                year=ano,
                month=mes,
                day=dia,
            ))
            imported += 1
        except Exception:
            skipped += 1

    db.session.commit()
    if imported:
        flash(f'{imported} despesa(s) importada(s) com sucesso! {skipped} linha(s) ignorada(s).', 'success')
    else:
        flash(f'Nenhuma despesa importada. {skipped} linha(s) ignorada(s). Verifique se os nomes dos membros correspondem aos do grupo atual.', 'warning')
    return redirect(url_for('expenses.index'))


# Redireciona rota antiga para nova
@expenses_bp.route('/import-c6', methods=['GET'])
def import_c6():
    return redirect(url_for('expenses.import_bank'))


@expenses_bp.route('/import/parse', methods=['POST'])
def import_bank_parse():
    users = tenant_users().order_by(User.name).all()
    uids = [u.id for u in users]

    user_id = request.form.get('user_id', type=int)
    if not user_id or user_id not in uids:
        flash('Selecione um usuário válido.', 'danger')
        return redirect(url_for('expenses.import_bank'))

    fmt    = request.form.get('format', '')
    bank   = request.form.get('bank', '').strip()
    upload = request.files.get('file')

    if not upload or not upload.filename:
        flash('Selecione um arquivo.', 'danger')
        return redirect(url_for('expenses.import_bank'))

    file_bytes = upload.read()

    parser = _PARSERS.get(fmt)
    if parser is None:
        flash('Formato não reconhecido.', 'danger')
        return redirect(url_for('expenses.import_bank'))

    try:
        transactions, bank_override = parser(file_bytes, request.form)
        if bank_override:
            bank = bank_override
    except Exception as e:
        msg = str(e).lower()
        if 'password' in msg or 'encrypt' in msg or 'decrypt' in msg:
            flash('O PDF está protegido por senha. Informe a senha correta.', 'danger')
        else:
            flash(f'Erro ao processar o arquivo: {e}', 'danger')
        return redirect(url_for('expenses.import_bank'))

    if not transactions:
        flash('Nenhuma transação de saída encontrada. Verifique o arquivo e o formato selecionado.', 'warning')
        return redirect(url_for('expenses.import_bank'))

    selected_user = next(u for u in users if u.id == user_id)
    all_uids = [u.id for u in users]
    tenant_id = request.args.get('_tid') or getattr(selected_user, 'tenant_id', None)
    if not tenant_id:
        from flask import session as _s
        tenant_id = _s.get('tenant_id')

    # Categorização automática (regras + histórico)
    cat_results = categorize_batch(
        [t.description for t in transactions], all_uids, tenant_id
    )
    suggested_cats = [r[0] for r in cat_results]
    cat_sources    = [r[1] for r in cat_results]

    # Detecção de duplicatas
    duplicates = set()
    for i, t in enumerate(transactions):
        if is_duplicate(user_id, t.year, t.month, t.day, t.amount, t.description):
            duplicates.add(i)

    return render_template(
        'expenses/import_preview.html',
        transactions=transactions,
        suggested_cats=suggested_cats,
        cat_sources=cat_sources,
        duplicates=duplicates,
        users=users,
        selected_user=selected_user,
        user_id=user_id,
        bank=bank,
        categories=CATEGORIES,
        total=sum(t.amount for t in transactions),
    )


@expenses_bp.route('/import/confirm', methods=['POST'])
def import_bank_confirm():
    uids = tenant_user_ids()
    indices = request.form.getlist('idx')
    count = 0

    for idx in indices:
        if not request.form.get(f'include_{idx}'):
            continue
        try:
            user_id        = int(request.form[f'user_id_{idx}'])
            if user_id not in uids:
                continue
            day            = int(request.form[f'day_{idx}'])
            month          = int(request.form[f'month_{idx}'])
            year           = int(request.form[f'year_{idx}'])
            if not (1 <= month <= 12 and 2000 <= year <= 2100 and 1 <= day <= 31):
                continue
            description    = request.form[f'desc_{idx}'].strip()
            amount         = float(request.form[f'amount_{idx}'].replace(',', '.'))
            category       = request.form.get(f'category_{idx}', 'Outros')
            payment_method = request.form.get(f'payment_method_{idx}', 'PIX')
            bank           = request.form.get(f'bank_{idx}', '')
        except (KeyError, ValueError):
            continue

        if not description or amount <= 0:
            continue

        db.session.add(Expense(
            user_id=user_id,
            description=description,
            amount=amount,
            category=category,
            payment_method=payment_method,
            bank=bank or None,
            year=year,
            month=month,
            day=day,
        ))
        count += 1

    db.session.commit()

    # Criar regras solicitadas pelo usuário na tela de preview
    from flask import session as _sess
    tid = _sess.get('tenant_id')
    if tid:
        rules_created = 0
        for idx in indices:
            if not request.form.get(f'criar_regra_{idx}'):
                continue
            desc = request.form.get(f'desc_{idx}', '').strip()
            cat  = request.form.get(f'category_{idx}', '').strip()
            # Usar a primeira palavra significativa como keyword
            words = [w for w in desc.upper().split() if len(w) >= 3]
            if words and cat:
                kw = words[0][:30]
                if create_rule_if_missing(kw, cat, tid):
                    rules_created += 1
        if rules_created:
            flash(f'{rules_created} regra(s) de categorização criada(s) automaticamente!', 'info')

    flash(f'{count} despesa(s) importada(s) com sucesso!', 'success')
    return redirect(url_for('expenses.index'))


def _write_expense_series(form, bank, payment, group_id, group_field, number_field, amounts, card_id=None):
    """Persiste N Expense rows mensais para um grupo já criado e flushado."""
    mes_inicio, ano_inicio = form.month.data, form.year.data
    for i, amount in enumerate(amounts):
        mes, ano = month_offset(mes_inicio, ano_inicio, i)
        db.session.add(Expense(
            user_id=form.user_id.data,
            description=form.description.data,
            amount=amount,
            category=form.category.data,
            payment_method=payment,
            bank=bank,
            year=ano,
            month=mes,
            day=form.day.data,
            card_id=card_id,
            **{group_field: group_id, number_field: i + 1},
        ))


def _create_installments(form, bank, card_id=None):
    n = form.num_installments.data
    total = Decimal(str(form.amount.data))
    parcela = (total / n).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    amounts = [parcela] * (n - 1) + [total - parcela * (n - 1)]

    group = InstallmentGroup(
        user_id=form.user_id.data,
        description=form.description.data,
        total_amount=total,
        num_installments=n,
        bank=bank or '',
    )
    db.session.add(group)
    db.session.flush()

    _write_expense_series(form, bank, 'Cartão de Crédito',
                          group.id, 'installment_group_id', 'installment_number', amounts,
                          card_id=card_id)
    db.session.commit()

    mes_inicio, ano_inicio = form.month.data, form.year.data
    mes_fim, ano_fim = month_offset(mes_inicio, ano_inicio, n - 1)
    parcela_fmt = _brl(parcela)
    flash(
        f'{n} parcelas criadas de {parcela_fmt} cada '
        f'({MONTH_NAMES_SHORT[mes_inicio-1]}/{ano_inicio} → {MONTH_NAMES_SHORT[mes_fim-1]}/{ano_fim})',
        'success'
    )


def _create_recurring(form, bank, payment, card_id=None):
    n = form.recurring_times.data
    amount = Decimal(str(form.amount.data))

    group = RecurringGroup(
        user_id=form.user_id.data,
        description=form.description.data,
        amount=amount,
        num_recurrences=n,
    )
    db.session.add(group)
    db.session.flush()

    _write_expense_series(form, bank, payment,
                          group.id, 'recurring_group_id', 'recurring_number', [amount] * n,
                          card_id=card_id)

    db.session.commit()

    mes_inicio, ano_inicio = form.month.data, form.year.data
    mes_fim, ano_fim = month_offset(mes_inicio, ano_inicio, n - 1)
    flash(
        f'Despesa recorrente criada por {n} meses '
        f'({MONTH_NAMES_SHORT[mes_inicio-1]}/{ano_inicio} → {MONTH_NAMES_SHORT[mes_fim-1]}/{ano_fim})',
        'success'
    )
